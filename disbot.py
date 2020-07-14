import os
import discord
import asyncio
import random
import openpyxl

client = discord.Client()

@client.event
async def on_ready():
	print('로그인 : '+(client.user.name))
	print('id : '+(client.user.id))
	print('------------------------')
	await client.change_presence(game=discord.Game(name='뺌 도움 <- 사용법', type=1))
	
@client.event
async def on_message(message):
	id = message.author.id

	if message.author.bot:
		return None
	if message.content.startswith('뺌 도움'):
		embed = discord.Embed(title="빼미에몽 사용법",
		description='뺌 골라 A B C..\n'
			'뺌 배워 <할말> <대답> - 말을 가르쳐\n'
			'뺌 <할말> - 빼미에몽 : <대답>\n'
            '뺌 감타디 - 감자타워디펜스 다운로드링크\n'
			'뺌 러시안 - 러시안룰렛 미니게임\n'
			'뺌 업다운 - 업다운 미니게임\n'
            '뺌 주사위(1~200)\n'
			'빼미에몽 - 대답을 해줘'
		,color=0x00ff00)
		await client.send_message(message.channel, embed=embed)
	elif message.content.startswith('뺌 골라'):
		choice = message.content.split(' ')
		choicenumber = random.randint(2, len(choice))
		choiceresult = choice[choicenumber]
		await client.send_message(message.channel, str(choiceresult)+'이(가) 좋겠네')
	elif message.content.startswith('뺌 감타디'):
		embed = discord.Embed(title="감자타워디펜스",
		description='다운로드 / 룰북 / 패치노트\nhttps://docs.google.com/spreadsheets/d/1xFTxiUZmJRuARkA7m-pHlYFo97Y0liyCdZbbPofazDM/edit?usp=sharing'
        ,color=0xff7f00)
		embed.set_footer(text= '윈도우에서 파일실행을 막을 수 있으나 문제없는 파일입니다.')
		await client.send_message(message.channel, embed=embed)
	elif message.content.startswith('빼미에몽님 배워주세요'):
		await client.send_message(message.channel, '하란다고 하네 ㅋㅋ')
	elif message.content.startswith('뺌 배워'):
		file = openpyxl.load_workbook('data.xlsx')
		sheet = file.active
		learn = message.content.split(' ')
		hmm = random.randint(1,13)
		if not hmm == 1:
			for i in range(1,257):
				if sheet['A'+str(i)].value == None or sheet['A'+str(i)].value == learn[2]:
					sheet['A'+str(i)].value = learn[2]
					for j in range(1,257):
						if sheet.cell(i,j).value == None:
							break
					sheet.cell(i,j).value = learn[3]
					await client.send_message(message.channel, '알았어 이제부터 '+str(learn[2])+'은(는) '+str(learn[3])+'이야')
					break
				if i == 256:
					await client.send_message(message.channel, '과부하! (최대 256단어)')
			file.save('data.xlsx')
		else:
			await client.send_message(message.channel, '싫은데? 빼미에몽님 배워주세요 라고 해봐')
	elif message.content.startswith('빼미에몽'):
		dosome = '왜,ㅖ,머,?,왜불러 할일이 그렇게 없어?'
		dosomechoice = dosome.split(',')
		dosomenumber = random.randint(1, len(dosomechoice)-1)
		dosomeresult = dosomechoice[dosomenumber]
		await client.send_message(message.channel, dosomeresult)
	elif message.content.startswith('뺌 러시안'):
		file = openpyxl.load_workbook('rr.xlsx')
		sheet = file.active
		sheet['A'+str(1)].value = random.randint(1,6)
		sheet['B'+str(1)].value = 0
		await client.send_message(message.channel, '러시안룰렛이 시작됬어 "뺌 당겨"로 쏴')
		file.save('rr.xlsx')
	elif message.content.startswith('뺌 당겨'):
		file = openpyxl.load_workbook('rr.xlsx')
		sheet = file.active
		if not sheet['A'+str(1)].value == 0:
			sheet['B'+str(1)].value += 1
			if sheet['A'+str(1)].value == sheet['B'+str(1)].value:
				sheet['A'+str(1)].value = 0
				sheet['B'+str(1)].value = 0
				await client.send_message(message.channel, '탕! <@'+id+'> 머리에 총맞았어? 머리아파?')
			else:
				await client.send_message(message.channel, '틱. '+str(sheet['B'+str(1)].value)+'번째는 통과')
			file.save('rr.xlsx')
		else:
			await client.send_message(message.channel, '뺌 러시안을 먼저해야지 멍청아')
	elif message.content.startswith('뺌 업다운'):
		file = openpyxl.load_workbook('rr.xlsx')
		sheet = file.active
		if sheet['A'+str(2)].value == 0:
			sheet['A'+str(2)].value = random.randint(1,100)
			sheet['B'+str(2)].value = 0
			await client.send_message(message.channel, '업다운을 시작했어 "업다운 1~100"로 할수있어 기회는 7회야')
		else:
			await client.send_message(message.channel, '아직 진행중이야')
		file.save('rr.xlsx')
	elif message.content.startswith('업다운'):
		file = openpyxl.load_workbook('rr.xlsx')
		sheet = file.active
		number = message.content.split(' ')
		if not int(sheet['A'+str(2)].value) == 0:
			sheet['B'+str(2)].value += 1
			if int(sheet['A'+str(2)].value) == int(number[1]):
				await client.send_message(message.channel, '어케맞췄노 시발련ㄴ아')
				sheet['A'+str(2)].value = 0
				sheet['B'+str(2)].value = 0
			else:
				if int(sheet['A'+str(2)].value) < int(number[1]):
					await client.send_message(message.channel, '다운 '+str(7-sheet['B'+str(2)].value)+'번 남았어')
				else:
					await client.send_message(message.channel, '업 '+str(7-sheet['B'+str(2)].value)+'번 남았어')
				if sheet['B'+str(2)].value == 7:
					await client.send_message(message.channel, '끝났네 답은 '+str(sheet['A'+str(2)].value)+'인데 멍청이')
			file.save('rr.xlsx')
    elif message.content.startswith('뺌 주사위'):
		await client.send_message(message.channel, '<@'+id+'>의 주사위 : '+random.randint(1,200))
	elif message.content.startswith('뺌'):
		file = openpyxl.load_workbook('data.xlsx')
		sheet = file.active
		memory = message.content.split(' ')
		for i in range(1,257):
			if sheet['A'+str(i)].value == memory[1]:
				words = []
				for j in range(1,257):
					if not sheet.cell(i,j+1).value == None:
						words.append(sheet.cell(i,j+1).value)
					else:
						answer = random.choice(words)
						await client.send_message(message.channel, str(answer))
						break
				break

access_token = os.environ["BOT_TOKEN"]
client.run(access_token)